"""
Sandia raw-xlsx loader (Arbin cycler export) for the SNL NMC/NCA cells.

The GitHub / Sandia raw release ships one Arbin .xlsx per cell-segment,
NOT the tidy BatteryArchive cycle_data.csv. Each .xlsx has a
"Channel_*" sheet with per-sample timeseries:

    Data_Point, Test_Time(s), Date_Time, Step_Time(s), Step_Index,
    Cycle_Index, Current(A), Voltage(V), Charge_Capacity(Ah),
    Discharge_Capacity(Ah), Charge_Energy(Wh), Discharge_Energy(Wh), ...

One physical cell is split across several dated files (e.g. NMC_1:
180227 -> 180507 -> 180605 -> ...). This loader:
  1. reads the Channel sheet of every file for a cell,
  2. for each Cycle_Index takes the max Discharge_Capacity(Ah) as that
     cycle's discharge capacity (last full-discharge value),
  3. drops capacity-check / RPT and partial cycles by keeping only
     physically plausible discharge capacities in [cap_lo, cap_hi],
  4. concatenates the per-file cycles in chronological (filename-date)
     order into a single monotone-index trajectory,
  5. computes SoH = capacity / Q_nom * 100 and returns the HIPIF schema:
       cell_id, cycle, cycle_norm, V_mean, V_min, V_max,
       I_mean, I_std, T_mean, T_max, SoH

Directory layout expected:
    {data_dir}/Sandia_raw/NMC/*.xlsx   (+ cell_list.xlsx anywhere)
    {data_dir}/Sandia_raw/NCA/*.xlsx

Temperature: these files carry no cell-temperature channel, so T_mean
is set from the cell's ambient set-point (parsed from the filename,
e.g. '35C') -> consistent with HIPIF's temperature-free target setting
(the value is masked during adaptation; used only if a constraint check
needs an ambient prior).
"""
from __future__ import annotations
import glob
import re
import numpy as np
import pandas as pd
from ..config import HIPIFConfig

NOMINAL_AH = {"NMC": 3.0, "NCA": 3.2}
CAP_BOUNDS = (1.5, 3.6)   # plausible per-cycle discharge Ah for a 3 Ah cell
_CHANNEL_RE = re.compile(r"^Channel_", re.I)


def _parse_date(fname):
    m = re.match(r"(\d{6})", fname.split("/")[-1])
    return m.group(1) if m else "000000"


def _parse_temp(fname):
    m = re.search(r"(\d{2})C", fname)
    return float(m.group(1)) if m else 25.0


def _parse_cellnum(fname):
    # NMC_1 / NMC1 -> 'a' ; NMC_2 / NMC2 -> 'b'
    m = re.search(r"(NMC|NCA)_?(\d)", fname, re.I)
    return m.group(2) if m else "1"


def _read_channel(path):
    """Return list of (cycle, dcap, v, i) sample rows from the Channel sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if _CHANNEL_RE.match(s)), None)
    if sheet is None:
        wb.close()
        return []
    ws = wb[sheet]
    idx = {}
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            idx = {h: j for j, h in enumerate(row) if h}
            ci = idx.get("Cycle_Index")
            di = idx.get("Discharge_Capacity(Ah)")
            vi = idx.get("Voltage(V)")
            ii = idx.get("Current(A)")
            if ci is None or di is None:
                wb.close()
                return []
            continue
        c = row[ci]
        if c is None:
            continue
        out.append((int(c),
                    row[di] if row[di] is not None else 0.0,
                    row[vi] if vi is not None and row[vi] is not None else np.nan,
                    row[ii] if ii is not None and row[ii] is not None else np.nan))
    wb.close()
    return out


def _aggregate_cell(files, cell_id, temp_C, nominal_ah, verbose):
    """Concatenate dated files for one cell -> per-cycle records."""
    cap_lo, cap_hi = CAP_BOUNDS
    records = []          # (global_cycle, cap, v_mean, v_min, v_max, i_mean, i_std)
    gcyc = 0
    for f in sorted(files, key=_parse_date):
        samples = _read_channel(f)
        if not samples:
            if verbose:
                print(f"  [Sandia-xlsx] {f.split('/')[-1]}: no Channel sheet, skipped")
            continue
        df = pd.DataFrame(samples, columns=["cyc", "dcap", "v", "i"])
        for cyc, g in df.groupby("cyc"):
            cap = float(g["dcap"].max())
            if not (cap_lo < cap < cap_hi):   # drop RPT / partial / zero
                continue
            gcyc += 1
            vv = g["v"].dropna()
            ii = g["i"].dropna()
            records.append({
                "cell_id": cell_id, "cycle": gcyc,
                "V_mean": float(vv.mean()) if len(vv) else 3.7,
                "V_min": float(vv.min()) if len(vv) else 3.2,
                "V_max": float(vv.max()) if len(vv) else 4.1,
                "I_mean": float(ii.mean()) if len(ii) else -1.0,
                "I_std": float(ii.std()) if len(ii) > 1 else 0.1,
                "T_mean": temp_C, "T_max": temp_C,
                "SoH": cap / nominal_ah * 100.0,
            })
    return records


def load_sandia_xlsx(cfg: HIPIFConfig, chemistry: str,
                     verbose: bool = True) -> pd.DataFrame:
    """chemistry in {'NMC','NCA'}. Reads {data_dir}/Sandia_raw/{chem}/*.xlsx."""
    assert chemistry in ("NMC", "NCA"), chemistry
    sdir = cfg.data_dir / "Sandia_raw" / chemistry
    files = [f for f in glob.glob(str(sdir / "*.xlsx"))
             if "cell_list" not in f.lower()]
    nominal = NOMINAL_AH[chemistry]

    if not files:
        raise FileNotFoundError(
            f"[Sandia-xlsx/{chemistry}] no .xlsx under {sdir} — synthetic "
            f"fallback is forbidden (P7); provide the raw data or drop the "
            f"NMC stage per decision D3")

    # group files by cell number (1/2 -> a/b)
    groups = {}
    for f in files:
        cn = _parse_cellnum(f)
        groups.setdefault(cn, []).append(f)

    all_rows = []
    for cn, fs in sorted(groups.items()):
        cell_id = f"SNL_18650_{chemistry}_{cn}"
        temp_C = _parse_temp(fs[0])
        rows = _aggregate_cell(fs, cell_id, temp_C, nominal, verbose)
        all_rows += rows
        if verbose and rows:
            s = pd.DataFrame(rows)
            print(f"[Sandia-xlsx/{chemistry}] {cell_id}: "
                  f"SoH [{s.SoH.min():.1f},{s.SoH.max():.1f}]%, "
                  f"n={len(rows)} cycles, T={temp_C:.0f}C")

    if not all_rows:
        raise RuntimeError(
            f"[Sandia-xlsx/{chemistry}] xlsx present but no usable cycles "
            f"parsed — inspect the raw files (fail-fast, P7)")

    df = pd.DataFrame(all_rows)
    for cid in df["cell_id"].unique():
        m = df["cell_id"] == cid
        mx = df.loc[m, "cycle"].max()
        df.loc[m, "cycle_norm"] = df.loc[m, "cycle"] / max(mx, 1)
    df["SoH"] = df["SoH"].clip(50, 105)
    return df.reset_index(drop=True)


def load_sandia_nmc_xlsx(cfg, verbose=True):
    return load_sandia_xlsx(cfg, "NMC", verbose)


def load_sandia_nca_xlsx(cfg, verbose=True):
    return load_sandia_xlsx(cfg, "NCA", verbose)


def _synthetic(cfg, chemistry, nominal):
    rng = np.random.default_rng(cfg.seed + (23 if chemistry == "NMC" else 29))
    rows = []
    base = 0.16 if chemistry == "NMC" else 0.20
    for k in range(2):
        cid = f"SYN_{chemistry}_{k+1}"
        n = rng.integers(150, 260); soh = 100.0
        for cyc in range(1, n + 1):
            T = 35 + rng.normal(0, 1.0)
            fade = base * (1.0 + (T - 35) / 60)
            if rng.random() < 0.04:
                fade -= 0.3
            soh = max(60.0, soh - fade)
            rows.append({
                "cell_id": cid, "cycle": cyc, "cycle_norm": cyc / n,
                "V_mean": 3.65 + rng.normal(0, 0.05),
                "V_min": 3.2, "V_max": 4.15,
                "I_mean": -1.0, "I_std": 0.1,
                "T_mean": T, "T_max": T + 3,
                "SoH": soh,
            })
    return pd.DataFrame(rows)
